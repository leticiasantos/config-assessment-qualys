import pandas as pd
from tkinter import Tk, filedialog
import os
from datetime import datetime
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

# Passo 1: Abrir o pop-up para selecionar o arquivo CSV
root = Tk()
root.withdraw()  # Esconder a janela principal

csv_filepath = filedialog.askopenfilename(
    title="Selecione o arquivo CSV",
    filetypes=[("CSV files", "*.csv")]
)

if not csv_filepath:
    print("Nenhum arquivo selecionado.")
    exit()

output_folder = os.path.dirname(csv_filepath)  # Pasta onde o CSV foi selecionado

# Passo 2: Efetuar a leitura do arquivo CSV e excluir as linhas antes de "RESULTS"
try:
    with open(csv_filepath, "r", newline='', encoding="utf-8") as csv_file:
        lines = csv_file.readlines()

    results_index = next((i for i, line in enumerate(lines) if "RESULTS" in line), None)
    if results_index is not None:
        lines = lines[results_index+1:]  # Pula a linha "RESULTS" também
    else:
        raise ValueError("Linha 'RESULTS' não encontrada.")
except Exception as e:
    print("Erro ao ler o arquivo CSV:", e)
    exit()

# Passo 3: Criar um DataFrame a partir dos dados após "RESULTS"
try:
    df = pd.read_csv(io.StringIO('\n'.join(lines)), sep=',', quotechar='"')
except Exception as e:
    print("Erro ao criar DataFrame:", e)
    exit()

# Passo 4: Excluir colunas não desejadas
columns_to_drop = [
    "DNS Hostname", "Qualys Host ID", "Tracking Method", "OS CPE", "NETWORK",
    "Last Scan Date", "Evaluation Date", "Operating System", "Criticality Value",
    "Instance", "Deprecated", "Qualys Host ID", "Cause of Failure"
]
df.drop(columns=columns_to_drop, inplace=True)

# Passo 5: Renomear colunas
if "NetBIOS Hostname" in df.columns:
    df.rename(columns={"NetBIOS Hostname": "Hostname"}, inplace=True)
if "Control ID" in df.columns:
    df.rename(columns={"Control ID": "Control ID Qualys"}, inplace=True)
if "Control References" in df.columns:
    df.rename(columns={"Control References": "Control ID CIS"}, inplace=True)
if "Criticality Label" in df.columns:
    df.rename(columns={"Criticality Label": "Criticality"}, inplace=True)

# Passo 6: Adicionar novas colunas
new_columns = ["Valor CIS", "Valor Identificado", "Valor Customizado", "Justificativa"]
for column in new_columns:
    df[column] = ""

# Passo 7: Limpar as strings "=====" da coluna "Evidence"
if "Evidence" in df.columns:
    df["Evidence"] = df["Evidence"].str.replace("======", "")
if "J" in df.columns:
    df["J"] = df["J"].str.replace(r"Current Value\(s\).*\(GMT-0300\)", "", regex=True)

# Passo 8: Adicionar fórmulas nas colunas "Valor CIS" e "Valor Identificado"
# Função para extrair Valor CIS
def formula_cis(evidence):
    if "Expected Value(s)" not in evidence:
        return ""
    else:
        start = evidence.find("Expected Value(s)") + len("Expected Value(s)")
        end = evidence.find("Current Value(s)") - evidence.find("Expected Value(s)") - len("Expected Value(s)")
        return evidence[start:start + end]

def formula_valor_identificado(evidence):
    if "Current Value(s)" in evidence:
        start_index = evidence.find("Current Value(s) ") + len("Current Value(s) ")
        end_index = evidence.find("(GMT-0300)") - evidence.find("Current Value(s)") - len("Current Value(s) ")
        formula_result = "=SUBSTITUTE(\"Current Value(s) \" & RIGHT(j:j, LEN(j:j) - FIND(\"(GMT-0300)\", J:J) - 11), \"Current Value(s)\", \"\")"
        return formula_result
    else:
        return ""
    
# Adicionar coluna "Valor Identificado"
df['Valor Identificado'] = df['Evidence'].apply(formula_valor_identificado)

# Adicionar coluna "Valor CIS"
df['Valor CIS'] = df['Evidence'].apply(formula_cis)

# Adicionar coluna "Valor Identificado"
df['Valor Identificado'] = df['Evidence'].apply(formula_valor_identificado)

# Substituir "------------ OR ------------" por "OR" na coluna K
df['Valor CIS'] = df['Valor CIS'].str.replace("------------ OR ------------", "OR")

# Substituir espaços em branco extras e caracteres de nova linha nas colunas K e L
df['Valor CIS'] = df['Valor CIS'].str.replace(r'\s+', ' ', regex=True).str.strip()
df['Valor Identificado'] = df['Valor Identificado'].str.replace(r'\s+', ' ', regex=True).str.strip()

# Passo 9: Salvar o DataFrame como um arquivo Excel e aplicar estilo de células
try:
    # Obter a data atual no formato YYYYMMDD
    current_date = datetime.now().strftime("%Y%m%d")

    # Pedir ao usuário para inserir informações
    nome_cliente = input("Digite o nome do cliente: ").strip()
    SO = input("Digite o sistema operacional: ").strip()
    version = input("Digite a versão da política: ").strip()

    # Construir o nome do arquivo
    output_name = f"{current_date}_Assessment_{nome_cliente}_{SO}_{version}"
    output_filename = f"{output_name}.xlsx"
    output_filepath = os.path.join(output_folder, output_filename)

    # Salvar o DataFrame como Excel
    df.to_excel(output_filepath, index=False)

    # Carregar o arquivo Excel e aplicar estilo
    wb = load_workbook(output_filepath)
    ws = wb.active
    fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")  # Cor aRGB é "AARRGGBB"
    font = Font(color="FFFFFF", bold=True)
    for cell in ws["1:1"]:
        cell.fill = fill
        cell.font = font

    # Ocultar a coluna J (Evidence)
    ws.column_dimensions['J'].hidden = True
    

    # Definir a largura de todas as colunas
    column_width_default = 17.00  # Largura desejada para as colunas padrão (17 unidades)
    column_width_centered = 13.00   # Largura desejada para as colunas centralizadas (12 unidades)

    for column in ws.columns:
        column_name = column[0].column_letter  # Letra da coluna (A, B, C, etc.)
        if column_name in ["A", "B", "C", "D", "E", "G", "H"]:
            ws.column_dimensions[column_name].width = column_width_centered
        else:
            ws.column_dimensions[column_name].width = column_width_default

    # Centralizar colunas específicas
    columns_to_center = ["A", "B", "C", "D", "E", "G", "H"]
    for col in columns_to_center:
        for cell in ws[f"{col}"]:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Centralizar colunas específicas
    columns_to_center = ["F", "I", "J", "K", "L", "M", "N"]
    for col in columns_to_center:
        for cell in ws[f"{col}"]:
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Definir quebra de linha, alinhamento central e vertical no cabeçalho das colunas de A até N
    for col in ws.iter_cols(min_col=1, max_col=14, max_row=1):
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    # Remover quebra de texto nas colunas K e L
    for col in ws.iter_cols(min_col=11, max_col=12):
        for cell in col:
            cell.alignment = Alignment(wrap_text=False)

    # Definir altura da linha (altura da célula)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):  # Começar da segunda linha (após cabeçalho)
        for cell in row:
            cell.font = Font(size=10)  # Definir tamanho da fonte
            ws.row_dimensions[cell.row].height = 34.50  # Definir altura da linha



    wb.save(output_filepath)
    print(f"Arquivo Excel atualizado salvo como '{output_filepath}'.")
except Exception as e:
    print("Erro ao salvar o arquivo Excel atualizado:", e)
